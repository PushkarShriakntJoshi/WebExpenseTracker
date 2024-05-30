import streamlit as st
import pandas as pd
from openpyxl import load_workbook, Workbook
from datetime import date
import matplotlib.pyplot as plt
from io import BytesIO
import os
import random

FILE_NAME = "expenses.xlsx"
one_liners = [
    "Money talks, but all mine ever says is goodbye.",
    "Why don’t they just print more money and give it to everyone?",
    "My wallet is like an onion—opening it makes me cry.",
    "I’m not great at math, but I can tell when I’m broke.",
    "I’m on a seafood diet. I see food and I buy it.",
    "A penny saved is ridiculous.",
    "If money doesn’t grow on trees, why do banks have branches?",
    "I used to be a banker, but I lost interest.",
    "I told my wife she should embrace her mistakes. She gave me a hug.",
    "Money is the root of all wealth.",
    "The safest way to double your money is to fold it over and put it in your pocket.",
    "I’m not broke; I’m just a pre-rich.",
    "Budget: a mathematical confirmation of your suspicions.",
    "A bank is a place that will lend you money if you can prove you don’t need it.",
    "I put a dollar in a change machine. Nothing changed.",
    "I can resist everything except temptation… and sales.",
    "I have a perfect record when it comes to saving money. I’ve never saved any.",
    "I’m having an out-of-money experience.",
    "Money can’t buy happiness, but it can buy a yacht big enough to pull up right alongside it.",
    "The quickest way to double your money is to fold it and put it back in your pocket.",
    "I’m so poor I can’t even pay attention.",
    "I used to be rich, but then my parents lost my allowance.",
    "I bought a vacuum cleaner. It’s been gathering dust ever since.",
    "Money is like a sixth sense—you can't make use of the other five without it.",
    "I'm living beyond my means, but I can afford it.",
    "Whoever said money can't buy happiness didn't know where to shop.",
    "I spent a lot of money on booze, birds, and fast cars. The rest I just squandered.",
    "I can’t afford to be a penny pincher.",
    "Money can't buy happiness, but it's a lot more comfortable to cry in a Mercedes than on a bicycle.",
    "I live paycheck to paycheck, but I have a really nice paycheck.",
    "Save money. Travel. Live well. Those are my goals. I’m 0 for 3 so far.",
    "I have a love-hate relationship with money. I love it; it hates me.",
    "I thought I wanted a career. Turns out, I just wanted paychecks.",
    "I wish I had a dollar for every dollar I don’t have.",
    "I got a credit card in my name. Now it’s my identity.",
    "A budget tells us what we can’t afford, but it doesn’t keep us from buying it.",
    "I don’t need a hair stylist. My pillow gives me a new hairstyle every morning.",
    "Money isn’t everything, but it sure keeps you in touch with your children.",
    "I’d give up chocolate, but I’m not a quitter.",
    "They say love is more important than money. Have you ever tried paying your bills with a hug?",
    "I love money, but it's always on the run.",
    "Always borrow money from a pessimist. He won’t expect it back.",
    "I can’t afford to waste my time anymore.",
    "I am not a millionaire, but I just know I’d be darling at it.",
    "Money is like a new toy. Once you lose it, you want it back.",
    "I want to live like a poor person with lots of money.",
    "I can’t believe how much money I spend on things I never use.",
    "I have enough money to last me the rest of my life, unless I buy something.",
    "I need a six-month vacation, twice a year.",
    "I’m so broke, I can’t even pay attention.",
    "I spend money like I’m rich. I just don’t make money like I’m rich.",
    "Why do banks charge you a 'non-sufficient funds fee' on money they already know you don’t have?",
    "I wish the bank would give me a loan, and then leave me alone.",
    "I spend money like there’s no tomorrow. Because, I’m broke by tomorrow.",
    "I’m great at saving money. I just don’t know where it is.",
    "I'm not broke, I'm just having a temporary cash flow issue.",
    "When I get money, I spend it. What else is it good for?",
    "I’m on a 30-day diet. So far, I’ve lost 15 days.",
    "Why is there so much month left at the end of the money?",
    "I have all the money I’ll ever need. If I die by 4 PM.",
    "I don't have an attitude problem. You have a perception problem.",
    "The lottery is a tax on people who are bad at math.",
    "My favorite machine at the gym is the vending machine.",
    "I’m in shape. Round is a shape, isn’t it?",
    "I’ll stop at nothing to avoid spending money.",
    "Money can’t buy happiness, but it can pay for therapy.",
    "I’m so poor, even my dreams are on layaway.",
    "I used to think I was indecisive, but now I’m not too sure.",
    "I’m so broke, my credit card bill is in Roman numerals.",
    "I have enough money to last me the rest of my life, unless I buy something.",
    "The best way to teach your kids about taxes is to eat 30% of their ice cream.",
    "Why is money called dough? Because we all knead it.",
    "I’m not cheap, I’m frugal.",
    "I’m saving for a rainy day, which is why I’m broke on sunny days.",
    "If you think nobody cares if you're alive, try missing a couple of car payments.",
    "I always arrive late at the office, but I make up for it by leaving early.",
    "I want to be a millionaire just like my uncle. He’s not a millionaire, but he wants to be one."
]

def init_excel():
    if not os.path.exists(FILE_NAME):
        wb = Workbook()
        ws = wb.active
        ws.append(["ID", "Expense Type", "Amount", "Date"])
        wb.save(FILE_NAME)

def load_expenses():
    wb = load_workbook(FILE_NAME)
    ws = wb.active
    data = ws.iter_rows(min_row=2, values_only=True)
    expenses = pd.DataFrame(data, columns=["ID", "Expense Type", "Amount", "Date"])
    return expenses

def add_expense(expense_type, amount, expense_date):
    wb = load_workbook(FILE_NAME)
    ws = wb.active
    next_id = ws.max_row  # Ensure next_id is always unique
    ws.append([next_id, expense_type, amount, expense_date])
    wb.save(FILE_NAME)

def delete_expense(expense_id):
    wb = load_workbook(FILE_NAME)
    ws = wb.active
    for row in ws.iter_rows(min_row=2):
        if row[0].value == expense_id:
            ws.delete_rows(row[0].row, 1)
            wb.save(FILE_NAME)
            return True
    return False

def plot_bar_chart(expenses):
    fig, ax = plt.subplots(figsize=(8, 5))
    expense_data = expenses.groupby("Expense Type")["Amount"].sum()
    ax.bar(expense_data.index, expense_data.values)
    ax.set_xlabel("Expense Type")
    ax.set_ylabel("Total Amount")
    ax.set_title("Monthly Expenses")
    plt.xticks(rotation=45)
    return fig

def export_to_excel(expenses):
    output = BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        expenses.to_excel(writer, index=False, sheet_name='Expenses')
        writer.save()  # Properly close the writer
    processed_data = output.getvalue()
    return processed_data

def show_one_liner():
    one_liner = random.choice(one_liners)
    st.markdown(f"<h3 style='color:green;'>{one_liner}</h3>", unsafe_allow_html=True)

init_excel()

st.title("Monthly Expense Tracker")

# Display a random one-liner on top of the screen
show_one_liner()

st.header("Add a New Expense")

expense_types = ["Grocery", "Fruits", "Vegetables", "Travel", "Fee", "Salary", "Online Shopping", "Misc"]

# Initialize the selected_expense_type variable
selected_expense_type = None

# Arrange buttons in two lines
cols = st.columns(4)
for i, expense_type in enumerate(expense_types):
    with cols[i % 4]:
        if st.button(expense_type):
            selected_expense_type = expense_type
            st.session_state["selected_expense_type"] = selected_expense_type

if "selected_expense_type" in st.session_state:
    selected_expense_type = st.session_state["selected_expense_type"]

amount = st.number_input("Amount", min_value=0.0, step=0.01)
expense_date = st.date_input("Date", value=date.today())

if st.button("Add Expense"):
    if selected_expense_type:
        add_expense(selected_expense_type, amount, expense_date)
        st.success("Expense added successfully!")
        # Show a new one-liner when expense is added
        show_one_liner()
    else:
        st.warning("Please select an expense type.")

st.header("Existing Expenses")

# Load the expenses and show only the last 5 entries
expenses = load_expenses()
if not expenses.empty:
    st.dataframe(expenses.tail(5))

if st.button("Export to Excel"):
    excel_data = export_to_excel(expenses)
    st.download_button(
        label="Download Excel",
        data=excel_data,
        file_name="expenses.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

st.header("Delete an Expense")

expense_id_to_delete = st.number_input("Expense ID to delete", min_value=1, step=1)
if st.button("Delete Expense"):
    if delete_expense(expense_id_to_delete):
        st.info("Expense deleted successfully.")
    else:
        st.warning("Expense ID not found.")

st.header("Monthly Expense Summary")

fig = plot_bar_chart(expenses)
st.pyplot(fig)
