if __name__ == "__main__":
    import streamlit as st
    import pandas as pd
    from openpyxl import load_workbook, Workbook
    from datetime import date
    import matplotlib.pyplot as plt
    from io import BytesIO
    import os

    FILE_NAME = "expenses.xlsx"

    def init_excel():
        if not os.path.exists(FILE_NAME):
            wb = Workbook()
            ws = wb.active
            ws.append(["ID", "Expense Type", "Amount", "Date"])
            wb.save(FILE_NAME)

    def load_expenses():
        wb = load_workbook(FILE_NAME)
        ws = wb.active
        data = ws.iter_rows(min_row=2, values_only=True)
        expenses = pd.DataFrame(data, columns=["ID", "Expense Type", "Amount", "Date"])
        return expenses

    def add_expense(expense_type, amount):
        wb = load_workbook(FILE_NAME)
        ws = wb.active
        next_id = ws.max_row
        today = date.today()
        ws.append([next_id, expense_type, amount, today])
        wb.save(FILE_NAME)

    def delete_expense(expense_id):
        wb = load_workbook(FILE_NAME)
        ws = wb.active
        for row in ws.iter_rows(min_row=2):
            if row[0].value == expense_id:
                ws.delete_rows(row[0].row, 1)
                wb.save(FILE_NAME)
                return True
        return False

    def plot_pie_chart(expenses):
        fig, ax = plt.subplots()
        expense_data = expenses.groupby("Expense Type")["Amount"].sum()
        ax.pie(expense_data, labels=expense_data.index, autopct='%1.1f%%')
        ax.set_title("Monthly Expenses")
        return fig

    init_excel()

    st.title("Monthly Expense Tracker")

    st.header("Add a New Expense")

    expense_types = ["Grocery", "Fruits", "Vegetables", "Travel", "Fee", "Salary", "Online Shopping", "Misc"]
    selected_expense_type = st.radio("Select Expense Type", expense_types)

    amount = st.number_input("Amount", min_value=0.0, format="%.2f")
    if st.button("Add Expense"):
        add_expense(selected_expense_type, amount)
        st.success("Expense added successfully")

    st.header("Delete an Expense")
    expense_id = st.number_input("Expense ID to delete", min_value=1)
    if st.button("Delete Expense"):
        if delete_expense(expense_id):
            st.success("Expense deleted successfully")
        else:
            st.error("Expense ID not found")

    st.header("Expenses")
    expenses = load_expenses()
    st.dataframe(expenses)

    st.header("Expense Chart")
    fig = plot_pie_chart(expenses)
    st.pyplot(fig)
